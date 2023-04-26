import openpyxl

# Setting Row Height and Column Width
wb = openpyxl.Workbook()
sheet = wb.active

sheet['A1'] = 'Tall row'
sheet['B2'] = 'Wide column'

# Set the height and width:
sheet.row_dimensions[1].height = 70
sheet.column_dimensions['B'].width = 20

wb.save('data/ch13/dimensions.xlsx')

# Merging and Unmerging Cells
wb = openpyxl.Workbook()
sheet = wb.active
sheet.merge_cells('A1:D3')  # Merge all these cells.
sheet['A1'] = 'Twelve cells merged together'

sheet.merge_cells('C5:D5')  # Merge these two cells.
sheet['C5'] = 'Two merged cells.'

wb.save('data/ch13/merged.xlsx')

openpyxl.load_workbook('data/ch13/merged.xlsx')

sheet = wb.active

sheet.unmerge_cells('A1:D3')  # Split these cells up.
sheet.unmerge_cells('C5:D5')

wb.save('data/ch13/merged.xlsx')

wb = openpyxl.load_workbook('data/ch13/produceSales.xlsx')
sheet = wb.active
sheet.freeze_panes = 'A2'  # Freeze the rows above A2.
wb.save('data/ch13/freezeExample.xlsx')
