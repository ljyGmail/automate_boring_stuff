import openpyxl

wb = openpyxl.Workbook()  # Create a blank workbook.
print(f'wb.sheetnames: {wb.sheetnames}')  # It starts with one sheet.

sheet = wb.active
print(f'sheet.title: {sheet.title}')

sheet.title = 'Spam Bacon Eggs Sheet'  # Change title.
print(f'wb.sheetnames: {wb.sheetnames}')

# Save a workbook.
wb = openpyxl.load_workbook('data/ch13/example.xlsx')
sheet = wb.active
sheet.title = 'Spam Spam Spam'
wb.save('data/ch13/example_copy.xlsx')

# Creating and Removing Sheets
wb = openpyxl.Workbook()
print(f'wb.sheetnames: {wb.sheetnames}')

print(f'wb.create_sheet(): {wb.create_sheet()}')  # Add a new sheet.
print(f'wb.sheetnames: {wb.sheetnames}')

print(
    f"wb.create_sheet(index=0, title='First Sheet'): {wb.create_sheet(index=0, title='First Sheet')}")  # Create a new sheet at index 0.
print(f'wb.sheetnames: {wb.sheetnames}')


print(
    f"wb.create_sheet(index=2,title='Middle Sheet'): {wb.create_sheet(index=2,title='Middle Sheet')}")
print(f'wb.sheetnames: {wb.sheetnames}')

del wb['Middle Sheet']
del wb['Sheet1']

print(f'wb.sheetnames: {wb.sheetnames}')

# Writing Values to Cells
sheet = wb['Sheet']
sheet['A1'] = 'Hello, world!'  # Edit the cell's value.
print(f"sheet['A1'].value: {sheet['A1'].value}")
