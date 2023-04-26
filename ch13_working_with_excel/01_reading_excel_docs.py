from openpyxl.utils import get_column_letter, column_index_from_string
import openpyxl

# Opening Excel Documents with OpenPyXL
wb = openpyxl.load_workbook('data/ch13/example.xlsx')
print(f'type(wb): {type(wb)}')

# Getting Sheets from the Workbook
print(f'wb.sheetnames: {wb.sheetnames}')  # The workbook's sheets' names.

sheet = wb['Sheet3']
print(f'sheet: {sheet}')
print(f'type(sheet): {type(sheet)}')
print(f'sheet.title: {sheet.title}')  # Get the sheet's title as a string.

anotherSheet = wb.active  # Get the active sheet.
print(f'anotherSheet: {anotherSheet}')

# Getting Cells from the Sheets
sheet = wb['Sheet1']  # Get a sheet from the workbook
print(f"sheet['A1']: {sheet['A1']}")  # Get a cell from the sheet.
print(f"sheet['A1'].value: {sheet['A1'].value}")

c = sheet['B1']
print(f"c.value: {c.value}")  # Get another cell from the sheet.

# Get the row, column, and value from the cell.
print('Row %s, Column %s is %s' % (c.row, c.column, c.value))
print('Cell %s is %s' % (c.coordinate, c.value))

print(f"sheet['C1'].value: {sheet['C1'].value}")

print(f'sheet.cell(row=1, column=2): {sheet.cell(row=1, column=2)}')
print(
    f'sheet.cell(row=1, column=2).value: {sheet.cell(row=1, column=2).value}')

for i in range(1, 8, 2):  # Go through every other row.
    print(i, sheet.cell(row=i, column=2).value)

# Determine the size of the sheet
sheet = wb['Sheet1']

print(f'sheet.max_row: {sheet.max_row}')  # Get the highest row number.


print(f'sheet.max_column: \
      {sheet.max_column}')  # Get the highest column number.

# Converting Between Column Letters and Numbers

print(f'get_column_letter(1): \
    {get_column_letter(1)}')  # Translate column 1 to a letter.
print(f'get_column_letter(2): {get_column_letter(2)}')
print(f'get_column_letter(27): {get_column_letter(27)}')
print(f'get_column_letter(900): {get_column_letter(900)}')

sheet = wb['Sheet1']
print(
    f'get_column_letter(sheet.max_column): {get_column_letter(sheet.max_column)}')


print(f"column_index_from_string('A'): \
    {column_index_from_string('A')}")  # Get A's number.

print(f"column_index_from_string('AA'): \
    {column_index_from_string('AA')}")


# Getting Rows and Columns from the Sheets
sheet = wb['Sheet1']

print(f"tuple(sheet['A1':'C3']): \
    {tuple(sheet['A1':'C3'])}")  # Get all cells from A1 to C3.

for rowOfCellObjects in sheet['A1':'C3']:
    for cellObj in rowOfCellObjects:
        print(cellObj.coordinate, cellObj.value)
    print('---END OF ROW---')

sheet = wb.active


print(f'list(sheet.columns)[1]: \
      {list(sheet.columns)[1]}')  # Get second column's cells

for cellObj in list(sheet.columns)[1]:
    print(cellObj.value)
